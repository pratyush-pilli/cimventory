import decimal
from django.db.models.aggregates import Sum, Min
from rest_framework.decorators import api_view, parser_classes
from rest_framework.response import Response
from rest_framework import status
from .models import DeliveryChallan, DeliveryChallanItem, InwardEntry, Inventory, ProjectCode, StockAllocation, StockOutward, LocationWiseAllocation, ReturnableGatePass, ReturnableGatePassItem, ReturnableGatePassReturn, ReturnableGatePassReturnItem, RejectedMaterialReturn, RejectedMaterialItem
from indent.models import Project, Requisition
from django.core.exceptions import ValidationError
from django.db import models
from decimal import Decimal
import traceback
from django.db import transaction
from django.utils import timezone
from datetime import datetime, timedelta
from django.conf import settings
import traceback
import logging
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import pdfkit
from rest_framework import viewsets
from store.serializers import ProjectCodeSerializer
from rest_framework import serializers
from purchase_order.models import PurchaseOrder, POLineItem
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
from django.http import HttpResponse
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.conf import settings
import json
from django.http import FileResponse
from django.core.paginator import Paginator
from .serializers import StockOutwardSerializer 
from users.models import CustomUser

# Performance optimization imports
import threading
from django.core.cache import cache
from django.db.models import Q, Prefetch
from django.core.cache.utils import make_template_fragment_key
from concurrent.futures import ThreadPoolExecutor

# File upload imports
from rest_framework.parsers import MultiPartParser, FormParser

# Add these imports at the top
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import permission_classes
from rest_framework.permissions import AllowAny

logger = logging.getLogger(__name__)

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def save_inward_entry(request):
    try:
        data = request.data
        logger.info(f"Received data: {data}")

        po_number = data.get('poNumber')
        received_date = data.get('receivedDate')
        location = data.get('location')
        remarks = data.get('remarks', "NA")
        items = data.get('items', [])
        
        # Handle file upload
        purchase_invoice = request.FILES.get('purchase_invoice')
        invoice_number = data.get('invoice_number')
        invoice_date = data.get('invoice_date')

        # Enhanced validation with better error messages
        if not po_number:
            return Response({'error': 'PO number is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not received_date:
            return Response({'error': 'Received date is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not location:
            return Response({'error': 'Location is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not items:
            return Response({'error': 'Items list cannot be empty'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate purchase invoice is required
        if not purchase_invoice:
            return Response({'error': 'Purchase invoice is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not invoice_number:
            return Response({'error': 'Invoice number is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not invoice_date:
            return Response({'error': 'Invoice date is required'}, status=status.HTTP_400_BAD_REQUEST)

        # OPTIMIZATION: Use select_related to reduce database queries
        try:
            po = PurchaseOrder.objects.select_related().get(po_number=po_number)
        except PurchaseOrder.DoesNotExist:
            return Response({'error': f'PO {po_number} not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check PO status
        if po.inward_status == 'completed':
            return Response({
                'error': 'This PO has been completely inwarded'
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Handle file upload first
            invoice_path = None
            if purchase_invoice:
                try:
                    # Create directory if it doesn't exist
                    media_root = os.path.join(settings.BASE_DIR, 'media')
                    invoices_dir = os.path.join(media_root, 'purchase_invoices')
                    os.makedirs(invoices_dir, exist_ok=True)
                    
                    # Create unique filename with PO number for mapping
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    file_extension = os.path.splitext(purchase_invoice.name)[1]
                    file_name = f"invoice_{po_number}_{timestamp}{file_extension}"  # PO number in filename
                    file_path = os.path.join(invoices_dir, file_name)
                    
                    # Save the file
                    with open(file_path, 'wb+') as destination:
                        for chunk in purchase_invoice.chunks():
                            destination.write(chunk)
                    
                    invoice_path = f'purchase_invoices/{file_name}'
                    logger.info(f"Purchase invoice saved for PO {po_number}: {invoice_path}")
                    
                except Exception as upload_error:
                    logger.error(f"Error saving purchase invoice for PO {po_number}: {str(upload_error)}")
                    return Response({
                        'error': 'Failed to save purchase invoice'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Parse items JSON if it's a string
            if isinstance(items, str):
                try:
                    items = json.loads(items)
                except json.JSONDecodeError:
                    return Response({'error': 'Invalid items format'}, status=status.HTTP_400_BAD_REQUEST)

            # OPTIMIZATION: Bulk fetch related objects to reduce queries
            item_codes = [item.get('itemCode') for item in items if item.get('itemCode')]
            po_line_items_map = {
                item.item_no: item 
                for item in POLineItem.objects.filter(
                    purchase_order=po,
                    item_no__in=item_codes
                ).select_related('purchase_order')
            }
            
            existing_inventories_map = {
                inv.item_no: inv 
                for inv in Inventory.objects.filter(item_no__in=item_codes)
            }
            
            # Batch processing lists
            inward_entries_batch = []
            po_items_to_update = []
            inventories_to_update = []
            new_inventories = []
            
            for item in items:
                item_code = item.get('itemCode')
                quantity_received = Decimal(str(item.get('quantityReceived', 0)))

                if not item_code:
                    raise ValidationError(f"Item code is required for item")
                
                if quantity_received <= 0:
                    continue  # Skip items with no quantity received

                if item_code not in po_line_items_map:
                    raise ValidationError(f"Item {item_code} not found in PO {po_number}")

                po_line_item = po_line_items_map[item_code]

                # Prepare inward entry for batch creation with invoice details
                # This creates the mapping between PO and invoice
                inward_entry = InwardEntry(
                    po_number=po_number,  # This links the invoice to the PO
                    received_date=received_date,
                    location=location,
                    remarks=remarks,
                    item_code=item_code,
                    description=item.get('description', "NA"),
                    make=item.get('make', "NA"),
                    material_group=item.get('material_group', "NA"),
                    ordered_quantity=item.get('orderedQuantity', 0),
                    quantity_received=quantity_received,
                    purchase_invoice=invoice_path,  # Invoice file path
                    invoice_number=invoice_number,  # Invoice details
                    invoice_date=invoice_date
                )
                inward_entries_batch.append(inward_entry)

                # Update PO line item quantities
                po_line_item.inwarded_quantity = (
                    po_line_item.inwarded_quantity or 0
                ) + quantity_received
                
                if po_line_item.inwarded_quantity > po_line_item.quantity:
                    raise ValidationError(
                        f"Inwarded quantity exceeds ordered quantity for item {item_code}"
                    )
                po_items_to_update.append(po_line_item)

                # Update or create inventory
                inventory = existing_inventories_map.get(item_code)
                
                if inventory:
                    # Update existing inventory
                    if location == "Times Square":
                        inventory.times_sq_stock += quantity_received
                    elif location == "iSquare":
                        inventory.i_sq_stock += quantity_received
                    elif location == "Sakar":
                        inventory.sakar_stock += quantity_received
                    elif location == "Pirana":
                        inventory.pirana_stock += quantity_received
                    else:
                        inventory.other_stock += quantity_received
                    
                    # Set inward_entry relationship after saving
                    inventories_to_update.append((inventory, inward_entry))
                else:
                    # Set location stock for new inventory
                    location_stock = {
                        'times_sq_stock': 0,
                        'i_sq_stock': 0,
                        'sakar_stock': 0,
                        'pirana_stock': 0,
                        'other_stock': 0
                    }
                    
                    if location == "Times Square":
                        location_stock['times_sq_stock'] = quantity_received
                    elif location == "iSquare":
                        location_stock['i_sq_stock'] = quantity_received
                    elif location == "Sakar":
                        location_stock['sakar_stock'] = quantity_received
                    elif location == "Pirana":
                        location_stock['pirana_stock'] = quantity_received
                    else:
                        location_stock['other_stock'] = quantity_received
                    
                    # Create new inventory
                    inventory = Inventory(
                        item_no=item_code,
                        description=item.get('description', "NA"),
                        make=item.get('make', "NA"),
                        material_group=item.get('material_group', "NA"),
                        **location_stock
                    )
                    new_inventories.append((inventory, inward_entry))

            # OPTIMIZATION: Perform bulk operations
            if inward_entries_batch:
                created_entries = InwardEntry.objects.bulk_create(inward_entries_batch)
                logger.info(f"Created {len(created_entries)} inward entries for PO {po_number}")
                
                # Update inventories with created inward entries
                for i, (inventory, _) in enumerate(inventories_to_update):
                    inventory.inward_entry = created_entries[i]
                    
                for i, (inventory, _) in enumerate(new_inventories):
                    inventory.save()  # Save first to get primary key
                    inventory.inward_entry = created_entries[len(inventories_to_update) + i]
                    inventory.save()  # Save again with inward_entry
            
            # Bulk update PO line items
            if po_items_to_update:
                POLineItem.objects.bulk_update(po_items_to_update, ['inwarded_quantity'])
            
            # Bulk update existing inventories
            if inventories_to_update:
                inventory_objects = [inv for inv, _ in inventories_to_update]
                Inventory.objects.bulk_update(
                    inventory_objects,
                    ['times_sq_stock', 'i_sq_stock', 'sakar_stock', 'pirana_stock', 'other_stock', 'inward_entry']
                )

            # Update PO status - CORRECTED LOGIC
            # Get all line items for this PO
            all_po_items = POLineItem.objects.filter(purchase_order=po)
            
            # Check status of all line items
            total_items = all_po_items.count()
            completed_items = 0
            partially_inwarded_items = 0
            
            for line_item in all_po_items:
                if line_item.inwarded_quantity >= line_item.quantity:
                    completed_items += 1
                elif line_item.inwarded_quantity > 0:
                    partially_inwarded_items += 1

            # Determine PO status based on all items
            if completed_items == total_items:
                po.inward_status = 'completed'
            elif completed_items > 0 or partially_inwarded_items > 0:
                po.inward_status = 'partially_inwarded'
            else:
                po.inward_status = 'open'

            # Update total inwarded quantity
            po.total_inwarded_quantity = sum(
                item.inwarded_quantity or 0 for item in all_po_items
            )
            
            po.save()
            logger.info(f"Updated PO {po_number} status to {po.inward_status} "
                   f"(Completed: {completed_items}/{total_items})")
                   
            # OPTIMIZATION: Get purchaser info efficiently
            purchaser = CustomUser.objects.filter(role='Purchaser').values('email', 'first_name', 'last_name').first()
            purchaser_email = purchaser['email'] if purchaser else None
            purchaser_first_name = purchaser['first_name'] if purchaser else 'Purchaser'
            purchaser_last_name = purchaser['last_name'] if purchaser else ''
            
            # OPTIMIZATION: Send email in background thread to avoid blocking response
            def send_email_background():
                try:
                    inward_email = InwardEmail()
                    inward_email.send('save_inward', purchaser_email=purchaser_email, purchaser_first_name=purchaser_first_name, purchaser_last_name=purchaser_last_name, po_number=po_number, items=items, po=po)
                except Exception as e:
                    logger.error(f"Error sending inward notification: {e}")
            
            # Start background email thread
            threading.Thread(target=send_email_background, daemon=True).start()

            return Response({
                'status': 'success',
                'message': 'Inward entry saved successfully',
                'po_status': po.inward_status,
                'invoice_path': invoice_path,
                'po_number': po_number,  # Return PO number for confirmation
                'invoice_number': invoice_number,  # Return invoice number for confirmation
                'details': {
                    'total_items': total_items,
                    'completed_items': completed_items,
                    'partially_inwarded': partially_inwarded_items
                }
            }, status=status.HTTP_201_CREATED)

    except ValidationError as e:
        logger.error(f"Validation error: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Error in save_inward_entry: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_inventory(request):
    try:
        # OPTIMIZATION: Add caching for inventory data (5 minutes)
        cache_key = 'inventory_list_data'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)
        
        # OPTIMIZATION: Use annotations to reduce database queries
        inventory_items = Inventory.objects.all().annotate(
            outward_stock=Sum('stockoutward__quantity')
        ).values(
            'id',
            'item_no',
            'description',
            'make',
            'material_group',
            'opening_stock',
            'allocated_stock',
            'available_stock',
            'times_sq_stock',
            'i_sq_stock',
            'sakar_stock',
            'pirana_stock',
            'other_stock',
            'total_stock',
            'outward_stock'
        )
        
        # Convert Decimal to float for JSON serialization
        processed_items = []
        for item in inventory_items:
            processed_item = {}
            for key, value in item.items():
                if isinstance(value, Decimal):
                    processed_item[key] = float(value or 0)
                else:
                    processed_item[key] = value
            processed_items.append(processed_item)

        # Cache the processed data for 5 minutes
        cache.set(cache_key, processed_items, 300)

        return Response(processed_items, status=status.HTTP_200_OK)
    except Exception as e:
        print("Error fetching inventory:", str(e))
        return Response(
            {'error': 'Failed to fetch inventory data.'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
def create_location_wise_allocation(allocation, location_allocations):
    """
    Create LocationWiseAllocation records for a given StockAllocation
    without modifying the location stock
    """
    location_mapping = {
        'Times Square': 'times_sq_stock',
        'iSquare': 'i_sq_stock',
        'Sakar': 'sakar_stock',
        'Pirana': 'pirana_stock',
        'Other': 'other_stock'
    }

    try:
        with transaction.atomic():
            for loc_alloc in location_allocations:
                display_location = loc_alloc.get('location', '').strip()
                quantity = Decimal(str(loc_alloc.get('quantity', 0)))
                
                if quantity <= 0 or not display_location:
                    continue

                db_location = location_mapping.get(display_location)
                if not db_location:
                    raise ValidationError(f"Invalid location: {display_location}")

                # Create allocation record without modifying location stock
                LocationWiseAllocation.objects.create(
                    stock_allocation=allocation,
                    location=db_location,
                    quantity=quantity
                )
                
                print(f"Created LocationWiseAllocation: {display_location} - {quantity}")
                
    except Exception as e:
        print(f"Error creating location-wise allocation: {str(e)}")
        raise e

@api_view(['POST'])
def allocate_stock(request):
    try:
        data = request.data
        print(f"Received allocation request: {data}")

        # Validate and convert inventory_id to integer
        try:
            inventory_id = int(data.get('inventory_id'))
        except (TypeError, ValueError):
            return Response({
                "message": "Invalid inventory ID format",
                "details": "inventory_id must be a number"
            }, status=status.HTTP_400_BAD_REQUEST)

        # OPTIMIZATION: Use select_related to get inventory
        try:
            inventory = Inventory.objects.select_related().get(id=inventory_id)
            print(f"Found inventory: {inventory.id}")

            # Calculate total location stock
            total_location_stock = (
                inventory.times_sq_stock +
                inventory.i_sq_stock +
                inventory.sakar_stock +
                inventory.pirana_stock +
                inventory.other_stock
            )
        except Inventory.DoesNotExist:
            return Response({
                "message": "Inventory not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Validate project_allocations is a list
        project_allocations = data.get('project_allocations', [])
        if not isinstance(project_allocations, list):
            return Response({
                "message": "Invalid project_allocations format",
                "details": "project_allocations must be an array"
            }, status=status.HTTP_400_BAD_REQUEST)

        # OPTIMIZATION: Batch fetch all project codes to reduce database queries
        project_codes = [str(entry.get("project_code", "")).strip() for entry in project_allocations if entry.get("project_code")]
        project_code_instances = {
            pc.code.project_code: pc 
            for pc in ProjectCode.objects.filter(
                code__project_code__in=project_codes
            ).select_related('code') if pc.code
        }

        # Process allocations within transaction
        with transaction.atomic():
            allocation_results = []
            
            for entry in project_allocations:
                # Validate project allocation entry
                if not isinstance(entry, dict):
                    return Response({
                        "message": "Invalid project allocation format",
                        "details": "Each allocation must be an object"
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Validate and clean project_code
                project_code = str(entry.get("project_code", "")).strip()
                if not project_code:
                    return Response({
                        "message": "Missing project code"
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Validate location_allocations
                location_allocations = entry.get("location_allocations", [])
                if not isinstance(location_allocations, list):
                    return Response({
                        "message": "Invalid location_allocations format",
                        "details": "location_allocations must be an array"
                    }, status=status.HTTP_400_BAD_REQUEST)

                # OPTIMIZATION: Use prefetched project code instance
                project_code_instance = project_code_instances.get(project_code)
                if not project_code_instance:
                    return Response({
                        "message": f"Project code {project_code} does not exist",
                        "code": "invalid_project_code"
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Calculate total quantity for this allocation
                try:
                    total_quantity = sum(
                        Decimal(str(loc.get('quantity', 0))) 
                        for loc in location_allocations
                    )
                except (TypeError, ValueError, decimal.InvalidOperation):
                    return Response({
                        "message": "Invalid quantity format in location allocations"
                    }, status=status.HTTP_400_BAD_REQUEST)

                if total_quantity <= 0:
                    return Response({
                        "message": "Total quantity must be greater than 0"
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Validate location-wise stock
                for loc_alloc in location_allocations:
                    location = loc_alloc.get('location', '').strip()
                    try:
                        quantity = Decimal(str(loc_alloc.get('quantity', 0)))
                    except (TypeError, ValueError, decimal.InvalidOperation):
                        return Response({
                            "message": f"Invalid quantity format for location {location}"
                        }, status=status.HTTP_400_BAD_REQUEST)

                    if quantity <= 0:
                        continue

                    # Map location name to database field
                    location_mapping = {
                        'Times Square': 'times_sq_stock',
                        'iSquare': 'i_sq_stock',
                        'Sakar': 'sakar_stock',
                        'Pirana': 'pirana_stock',
                        'Other': 'other_stock'
                    }
                    
                    db_location = location_mapping.get(location)
                    if not db_location:
                        return Response({
                            "message": f"Invalid location: {location}",
                            "valid_locations": list(location_mapping.keys())
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Check if location has enough stock
                    current_stock = getattr(inventory, db_location, Decimal('0'))
                    if current_stock < quantity:
                        return Response({
                            "message": f"Insufficient stock at {location}",
                            "details": {
                                "location": location,
                                "available": float(current_stock),
                                "requested": float(quantity)
                            }
                        }, status=status.HTTP_400_BAD_REQUEST)

                # Create allocation
                try:
                    allocation_date = data.get('allocation_date')
                    if allocation_date:
                        from dateutil.parser import parse
                        allocation_date = parse(allocation_date)
                except Exception as e:
                    return Response({
                        "message": "Invalid allocation_date format",
                        "details": str(e)
                    }, status=status.HTTP_400_BAD_REQUEST)

                allocation = StockAllocation.objects.create(
                    inventory=inventory,
                    project_code=project_code_instance,
                    allocated_quantity=total_quantity,
                    remarks=str(data.get('remarks', '')),
                    status='allocated',
                    allocation_date=allocation_date
                )

                # Create location-wise allocations
                create_location_wise_allocation(allocation, location_allocations)

                allocation_results.append({
                    "project_code": project_code,
                    "allocated_quantity": float(total_quantity),
                    "allocation_id": allocation.id
                })

            # Calculate final stock values - FIX HERE
            # Don't call calculate_available_stock() directly, update the values manually
            # to avoid the recursion issue
            total_allocated = StockAllocation.objects.filter(
                inventory=inventory,
                status='allocated'
            ).aggregate(total=Sum('allocated_quantity'))['total'] or 0
            
            inventory.allocated_stock = total_allocated
            inventory.available_stock = total_location_stock - total_allocated - inventory.outward_stock
            inventory.total_stock = total_location_stock
            inventory.save(update_fields=['allocated_stock', 'available_stock', 'total_stock'])

            # OPTIMIZATION: Clear inventory cache after allocation
            cache.delete('inventory_list_data')

            # Get updated values
            current_allocated = inventory.allocated_stock
            current_available = inventory.available_stock

            return Response({
                "message": "Stock allocated successfully",
                "allocations": allocation_results,
                "stock_status": {
                    "total_location_stock": float(total_location_stock),
                    "allocated_stock": float(current_allocated),
                    "available_stock": float(current_available),
                    "location_details": {
                        "times_square": float(inventory.times_sq_stock),
                        "i_square": float(inventory.i_sq_stock),
                        "sakar": float(inventory.sakar_stock),
                        "pirana": float(inventory.pirana_stock),
                        "other": float(inventory.other_stock)
                    }
                }
            }, status=status.HTTP_201_CREATED)

    except Exception as e:
        print(f"Error in allocate_stock: {str(e)}")
        traceback.print_exc()
        return Response({
            "message": "Failed to allocate stock",
            "error": str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_allocations(request):
    try:
        # Get inventory_id filter if provided
        inventory_id = request.GET.get('inventory_id')
        
        # OPTIMIZATION: Build base query with prefetch_related to reduce database queries
        query = StockAllocation.objects.select_related(
            'inventory', 
            'project_code',
            'project_code__code'
        ).prefetch_related(
            Prefetch(
                'location_allocations',
                queryset=LocationWiseAllocation.objects.all()
            )
        )
        
        # Apply filter if inventory_id is provided
        if inventory_id:
            query = query.filter(inventory_id=inventory_id)
        
        # Get all allocations
        allocations = query.all()
        
        allocation_data = []
        for allocation in allocations:
            # Handle the case where project_code or code might be None
            project_code_value = None
            project_name = None
            
            if allocation.project_code:
                if allocation.project_code.code:
                    project_code_value = allocation.project_code.code.project_code
                project_name = allocation.project_code.name
            
            # OPTIMIZATION: Use prefetched location-wise allocations
            location_allocation_data = []
            for loc_alloc in allocation.location_allocations.all():
                location_allocation_data.append({
                    'location': loc_alloc.location,
                    'quantity': float(loc_alloc.quantity)
                })
            
            allocation_data.append({
                'id': allocation.id,
                'inventory_id': allocation.inventory_id,
                'item_no': allocation.inventory.item_no,
                'description': allocation.inventory.description,
                'make': allocation.inventory.make,
                'material_group': allocation.inventory.material_group,
                'project_code': {
                    'code': project_code_value,
                    'name': project_name,
                },
                'allocated_quantity': float(allocation.allocated_quantity),
                'allocation_date': allocation.allocation_date,
                'remarks': allocation.remarks,
                'available_stock': float(allocation.inventory.available_stock),
                'status': allocation.status,
                'location_allocations': location_allocation_data
            })
            
        return Response(allocation_data)
        
    except Exception as e:
        # Enhance error logging
        print(f"Error in get_allocations: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def outward_stock(request):
    try:
        data = request.data
        print(f"Received outward request: {data}")

        outward_items = data.get('outward_items', [])
        if not outward_items:
            return Response({
                'error': 'No outward items provided'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Location display name to database field mapping
        LOCATION_MAPPING = {
            'Times Square': 'times_sq_stock',
            'iSquare': 'i_sq_stock',
            'Sakar': 'sakar_stock',
            'Pirana': 'pirana_stock',
            'Other': 'other_stock'
        }

        outward_records = []
        with transaction.atomic():
            for item in outward_items:
                inventory_id = item.get('inventory_id')
                project_code = item.get('project_code')
                location_quantities = item.get('location_quantities', {})
                outward_type = item.get('outward_type', 'available')  # Default to available if not specified
                
                if not all([inventory_id, project_code, location_quantities]):
                    raise ValidationError("Missing required item data")

                try:
                    inventory = Inventory.objects.get(id=inventory_id)
                    project_code_instance = ProjectCode.objects.get(code=project_code)
                except (Inventory.DoesNotExist, ProjectCode.DoesNotExist) as e:
                    raise ValidationError(f"Invalid inventory or project code: {str(e)}")
            
                for display_location, qty in location_quantities.items():
                    if not qty or float(qty) <= 0:
                        continue

                    db_location = LOCATION_MAPPING.get(display_location)
                    if not db_location:
                        raise ValidationError(f"Invalid location: {display_location}")

                    qty_decimal = Decimal(str(qty))
                    
                    # Validate stock availability at the location level
                    current_stock = getattr(inventory, db_location, 0)
                    if current_stock < qty_decimal:
                        raise ValidationError(
                            f"Insufficient stock at {display_location}: "
                            f"Available {current_stock}, Requested {qty_decimal}"
                        )

                    # Handle allocation differently based on outward_type
                    allocation = None
                    if outward_type == 'allocated':
                        # Check for allocation
                        location_allocation = LocationWiseAllocation.objects.filter(
                            stock_allocation__inventory=inventory,
                            stock_allocation__project_code=project_code_instance,
                            stock_allocation__status='allocated',
                            location=db_location
                        ).first()

                        if not location_allocation:
                            # Instead of failing, let's switch to 'available' with a warning
                            print(f"Warning: No allocation found for {display_location} for project {project_code}. Using available stock.")
                            # You could log this or track it in some way
                            outward_type = 'available'
                        elif location_allocation.quantity < qty_decimal:
                            raise ValidationError(
                                f"Insufficient allocated stock at {display_location}: "
                                f"Allocated {location_allocation.quantity}, Requested {qty_decimal}"
                            )
                        else:
                            allocation = location_allocation.stock_allocation

                    # Create outward entry
                    outward = StockOutward.objects.create(
                        inventory=inventory,
                        project_code=project_code_instance,
                        quantity=qty_decimal,
                        location=db_location,
                        outward_type=outward_type,  # Use the potentially modified outward_type
                        document_type=data.get('document_type'),
                        document_number=data.get('document_number'),
                        remarks=data.get('remarks', ''),
                        status='submitted',
                        stock_allocation=allocation  # This might be None if using available stock
                    )
                    outward_records.append(outward)
                    
                    # Update inventory stock
                    setattr(inventory, db_location, current_stock - qty_decimal)
                    
                    # Update allocation if needed
                    if outward_type == 'allocated' and allocation:
                        # Update location allocation
                        location_allocation.quantity -= qty_decimal
                        location_allocation.save()

                        # Update main allocation
                        allocation.allocated_quantity -= qty_decimal
                        if allocation.allocated_quantity <= 0:
                            allocation.status = 'fully_outward'
                        else:
                            allocation.status = 'partially_outward'
                        allocation.save()

                # Save inventory after processing all locations for this item
                inventory.calculate_available_stock()
                inventory.save()

            return Response({
                'success': True,
                'message': 'Outward processed successfully',
                'outward_ids': [outward.id for outward in outward_records]
            }, status=status.HTTP_201_CREATED)

    except ValidationError as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        print(f"Error in outward_stock: {str(e)}")
        traceback.print_exc()
        return Response({
            'error': 'Failed to process outward',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_project_requirements(request, project_code):
    try:
        project = Project.objects.get(project_code=project_code)
        required_items = project.get_required_items()
        
        # Aggregate quantities by item number
        aggregated_items = {}
        for item in required_items:
            item_no = item['cimcon_part_number']
            if item_no in aggregated_items:
                aggregated_items[item_no]['req_qty'] += item['req_qty']
            else:
                aggregated_items[item_no] = {
                    'cimcon_part_number': item_no,
                    'material_description': item['material_description'],
                    'make': item['make'],
                    'material_group': item['material_group'],
                    'req_qty': item['req_qty']
                }

        inventory_status = []
        for item in aggregated_items.values():
            inventory = Inventory.objects.filter(
                item_no=item['cimcon_part_number']
            ).first()
            
            if inventory:
                # Get total allocation
                allocation = StockAllocation.objects.filter(
                    inventory=inventory,
                    project_code__code=project_code
                ).aggregate(
                    total_allocated=models.Sum('allocated_quantity')
                )['total_allocated'] or 0
                
                # Get total outward
                outward = StockOutward.objects.filter(
                    inventory=inventory,
                    project_code=project_code
                ).aggregate(
                    total_outward=models.Sum('quantity')
                )['total_outward'] or 0
                
                inventory_status.append({
                    'item_no': item['cimcon_part_number'],
                    'description': item['material_description'],
                    'make': item['make'],
                    'material_group': item['material_group'],
                    'required_quantity': item['req_qty'],
                    'allocated_quantity': inventory.allocated_stock,
                    'outward_quantity': outward,
                    'pending_quantity': item['req_qty'] - outward,
                    'available_stock': inventory.available_stock,
                    'inventory_id': inventory.id,
                    'fifo_details': get_fifo_requisitions(project_code, item['cimcon_part_number'])
                })

        return Response(inventory_status)
        
    except Project.DoesNotExist:
        return Response(
            {'error': 'Project not found'},
            status=status.HTTP_404_NOT_FOUND
        )

def get_fifo_requisitions(project_code, item_no):
    """Get requisitions in FIFO order for a specific item"""
    return Requisition.objects.filter(
        project__project_code=project_code,
        cimcon_part_number=item_no,
        approved_status=True
    ).order_by('requisition_date').values(
        'id', 'req_qty', 'requisition_date'
    )

@api_view(['POST'])
def reallocate_stock(request, allocation_id):
    try:
        data = request.data
        print(f"Reallocation request data: {data}")
        
        # Get the original allocation
        allocation = StockAllocation.objects.get(id=allocation_id)
        print(f"Found original allocation: {allocation}")
        
        # Get the location-wise allocation
        location_allocation = LocationWiseAllocation.objects.filter(
            stock_allocation=allocation,
            location=data['location']
        ).first()
        print(f"Found location allocation: {location_allocation}")
        
        if not location_allocation:
            return Response({
                'error': f'No allocation found for project {allocation.project_code.code} and location {data["location"]}'
            }, status=status.HTTP_404_NOT_FOUND)
            
        # Get project code from either new_project_code or targetProject
        project_code = data.get('new_project_code') or data.get('targetProject')
        print(f"Looking for project code: {project_code}")
        
        # Debug: List all available project codes
        all_codes = ProjectCode.objects.all().values_list('code', flat=True)
        print(f"Available project codes: {list(all_codes)}")
        
        if not project_code:
            return Response({
                'error': 'Project code is required'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        # Validate new project code
        try:
            new_project = ProjectCode.objects.get(code=project_code)
            print(f"Found project: {new_project}")
        except ProjectCode.DoesNotExist:
            return Response({
                'error': f'Invalid project code: {project_code}. Available codes: {list(all_codes)}'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        # Validate quantity
        quantity = Decimal(str(data['quantity']))
        if quantity <= 0 or quantity > location_allocation.quantity:
            return Response({
                'error': 'Invalid quantity',
                'max_available': float(location_allocation.quantity),
                'requested': float(quantity)
            }, status=status.HTTP_400_BAD_REQUEST)
            
        with transaction.atomic():
            # Create new allocation for target project
            new_allocation = StockAllocation.objects.create(
                inventory=allocation.inventory,
                project_code=new_project,
                allocated_quantity=quantity,
                remarks=data.get('remarks', ''),
                status='allocated'
            )
            
            # Create new location-wise allocation
            LocationWiseAllocation.objects.create(
                stock_allocation=new_allocation,
                location=data['location'],
                quantity=quantity
            )
            
            # Update original allocation
            location_allocation.quantity -= quantity
            location_allocation.save()
            
            allocation.allocated_quantity -= quantity
            if allocation.allocated_quantity <= 0:
                allocation.status = 'reallocated'
            allocation.save()
            
        return Response({
            'message': 'Stock reallocated successfully',
            'new_allocation_id': new_allocation.id
        })
        
    except Exception as e:
        print(f"Error in reallocate_stock: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_outward_history(request, inventory_id, project_code):
    """
    Get the total outwarded quantity for a specific inventory item and project
    """
    try:
        # Get ProjectCode instance for the provided project code
        project_code_obj = ProjectCode.objects.get(code=project_code)
        
        # Calculate total outwarded quantity
        total_outwarded = StockOutward.objects.filter(
            inventory_id=inventory_id,
            project_code_id=project_code_obj.id  # Use project_code_id to match your data structure
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        # Get the most recent outward transactions (optional)
        recent_outwards = StockOutward.objects.filter(
            inventory_id=inventory_id,
            project_code_id=project_code_obj.id
        ).order_by('-created_at')[:5].values(
            'quantity', 
            'location', 
            'created_at', 
            'document_number', 
            'document_type', 
            'outward_type'
        )
        
        return Response({
            'total_outwarded': total_outwarded,
            'recent_outwards': recent_outwards
        })
    except ProjectCode.DoesNotExist:
        return Response({'error': f'Project code {project_code} not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def get_stock_details(request, item_id, project_code):
    try:
        inventory = Inventory.objects.get(id=item_id)
        
        # Get allocated stock for this project
        allocated_stock = StockAllocation.objects.filter(
            inventory=inventory,
            project_code__code=project_code
        ).aggregate(
            total=models.Sum('allocated_quantity')
        )['total'] or 0
        
        # Get total outwarded for this project
        total_outwarded = StockOutward.objects.filter(
            inventory=inventory,
            project_code=project_code
        ).aggregate(
            total=models.Sum('quantity')
        )['total'] or 0
        
        return Response({
            'allocated_stock': float(allocated_stock),
            'available_stock': float(inventory.available_stock),
            'total_outwarded': float(total_outwarded),
            'times_sq_stock': float(inventory.times_sq_stock),
            'i_sq_stock': float(inventory.i_sq_stock),
            'sakar_stock': float(inventory.sakar_stock),
            'pirana_stock': float(inventory.pirana_stock),
            'other_stock': float(inventory.other_stock),
        })
        
    except Inventory.DoesNotExist:
        return Response(
            {'error': 'Inventory not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
def generate_document(request):
    try:
        data = request.data
        logger.info(f"Received document generation request: {data}")

        # Validate outward items have HSN code and rate
        for item in data.get('outward_items', []):
            if not item.get('hsn_code'):
                return Response({
                    'error': 'HSN/SAC code is required for all items',
                    'item': item.get('item_no')
                }, status=400)
            
            if not item.get('rate'):
                return Response({
                    'error': 'Rate is required for all items',
                    'item': item.get('item_no')
                }, status=400)

        # Get project code - handle both projectCode and project_code
        project_code = data.get('project_code')
        if not project_code:
            # Try alternate field name if project_code is not found
            project_code = data.get('projectCode')
            
        if not project_code:
            return Response({
                'error': 'Project code is required'
            }, status=400)
            
        # Get project code instance
        try:
            project_code_obj = ProjectCode.objects.get(code__project_code=project_code)
        except ProjectCode.DoesNotExist:
            return Response({
                'error': f'Project code {project_code} not found'
            }, status=404)

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Challan"

        # Set column widths
        ws.column_dimensions['A'].width = 5   # Sr.No.
        ws.column_dimensions['B'].width = 15  # Item No.
        ws.column_dimensions['C'].width = 40  # Description
        ws.column_dimensions['D'].width = 15  # Make
        ws.column_dimensions['E'].width = 15  # Material Group
        ws.column_dimensions['F'].width = 15  # HSN/SAC
        ws.column_dimensions['G'].width = 10  # Quantity
        ws.column_dimensions['H'].width = 10  # Rate
        ws.column_dimensions['I'].width = 15  # Amount

        # Document Header
        ws.merge_cells('A1:I1')
        ws['A1'] = "CIMCON SOFTWARE INDIA PRIVATE LIMITED"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Address
        ws.merge_cells('A2:I2')
        ws['A2'] = "CIMCON Software India Pvt. Ltd., Plot No. 1, Sector 16A, Noida - 201301"
        ws['A2'].alignment = Alignment(horizontal='center')

        # GSTIN
        ws.merge_cells('A3:I3')
        ws['A3'] = "GSTIN: 09AABCC1234A1Z5"
        ws['A3'].alignment = Alignment(horizontal='center')

        # Document Details
        ws['A4'] = "Document No.:"
        ws['B4'] = data.get('documentNumber')
        ws['D4'] = "Date:"
        ws['E4'] = data.get('date')
        ws['G4'] = "Ref No.:"
        ws['H4'] = data.get('reference_no')

        # Project Details
        ws['A5'] = "Project Code:"
        ws['B5'] = data.get('projectCode')
        ws['D5'] = "Mode of Transport:"
        ws['E5'] = data.get('mode_of_transport')
        ws['G5'] = "Vehicle No.:"
        ws['H5'] = data.get('vehicle_no')

        # Additional Details
        ws['A6'] = "Supply Date:"
        ws['B6'] = data.get('supply_date')
        ws['D6'] = "Dispatch From:"
        ws['E6'] = data.get('dispatch_from')
        ws['G6'] = "Place of Supply:"
        ws['H6'] = data.get('place_of_supply')

        # Billing and Shipping Address
        ws['A7'] = "Billing Address:"
        ws.merge_cells('A8:C10')
        ws['A8'] = data.get('bill_to', '')
        ws['A8'].alignment = Alignment(wrap_text=True, vertical='top')

        ws['F7'] = "Shipping Address:"
        ws.merge_cells('F8:I10')
        ws['F8'] = data.get('ship_to', '')
        ws['F8'].alignment = Alignment(wrap_text=True, vertical='top')

        # Tax Details
        ws['A12'] = "CGST (%)"
        ws['B12'] = data.get('cgst', 0)
        ws['D12'] = "SGST (%)"
        ws['E12'] = data.get('sgst', 0)
        ws['G12'] = "IGST (%)"
        ws['H12'] = data.get('igst', 0)

        # Items Table Header
        headers = ['Sr.No.', 'Item No.', 'Description', 'Make', 'Material Group', 'HSN/SAC', 'Quantity', 'Rate', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=14, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

        # Items Data
        current_row = 15
        total_amount = 0
        challan_items = []
        
        for idx, item in enumerate(data.get('outward_items', []), 1):
            # Get the quantity from location quantities
            quantity = float(sum(item.get('location_quantities', {}).values()))
            
            # Get rate and calculate amount
            rate = float(item.get('rate', 0) or 0)
            amount = quantity * rate
            total_amount += amount

            # Ensure HSN code and rate are properly formatted
            hsn_code = item.get('hsn_code', '')  # Default to empty string if missing
            rate = float(item.get('rate', 0) or 0)  # Default to 0 if missing or invalid
            
            row_data = [
                idx,
                item.get('item_no', ''),
                item.get('description', ''),
                item.get('make', ''),
                item.get('material_group', ''),
                hsn_code,  # HSN code
                quantity,
                rate,     # Rate
                amount
            ]
            
            # Store item data for database
            challan_items.append({
                'inventory_id': item.get('inventory_id'),
                'item_no': item.get('item_no', ''),
                'description': item.get('description', ''),
                'make': item.get('make', ''),
                'material_group': item.get('material_group', ''),
                'hsn_code': hsn_code,
                'quantity': quantity,
                'rate': rate,
                'amount': amount,
                'uom': item.get('uom', 'NOS')
            })

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                if col in [8, 9]:  # Rate and Amount columns
                    cell.number_format = '#,##0.00'
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

            current_row += 1

        # Calculate and add tax amounts
        current_row += 1
        ws.cell(row=current_row, column=7).value = "Sub Total:"
        ws.cell(row=current_row, column=9).value = total_amount
        ws.cell(row=current_row, column=9).number_format = '#,##0.00'

        cgst_rate = float(data.get('cgst', 0) or 0)
        cgst_amount = total_amount * cgst_rate / 100
        current_row += 1
        ws.cell(row=current_row, column=7).value = f"CGST ({cgst_rate}%):"
        ws.cell(row=current_row, column=9).value = cgst_amount
        ws.cell(row=current_row, column=9).number_format = '#,##0.00'

        sgst_rate = float(data.get('sgst', 0) or 0)
        sgst_amount = total_amount * sgst_rate / 100
        current_row += 1
        ws.cell(row=current_row, column=7).value = f"SGST ({sgst_rate}%):"
        ws.cell(row=current_row, column=9).value = sgst_amount
        ws.cell(row=current_row, column=9).number_format = '#,##0.00'

        igst_rate = float(data.get('igst', 0) or 0)
        igst_amount = total_amount * igst_rate / 100
        current_row += 1
        ws.cell(row=current_row, column=7).value = f"IGST ({igst_rate}%):"
        ws.cell(row=current_row, column=9).value = igst_amount
        ws.cell(row=current_row, column=9).number_format = '#,##0.00'

        final_amount = total_amount + cgst_amount + sgst_amount + igst_amount
        current_row += 1
        ws.cell(row=current_row, column=7).value = "Total Amount:"
        ws.cell(row=current_row, column=9).value = final_amount
        ws.cell(row=current_row, column=9).number_format = '#,##0.00'

        # Add remarks if any
        if data.get('remarks'):
            current_row += 2
            ws.merge_cells(f'A{current_row}:I{current_row}')
            ws[f'A{current_row}'] = f"Remarks: {data.get('remarks')}"

        # Add note section
        current_row += 2
        ws.merge_cells(f'A{current_row}:I{current_row+2}')
        ws[f'A{current_row}'] = ("NOTE: Kindly check material in Good condition before taking delivery from Transporter. "
                                "If any Damages occurred then revert immediately within 1 Working Day. "
                                "There will not be acceptance of any damage complaint after 2 days of Receiving Material. "
                                "Contact your Sales Person Immediately If any Damages Occurred")
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)

        # Save the generated document
        media_root = os.path.join(settings.BASE_DIR, 'media')
        documents_dir = os.path.join(media_root, 'documents')
        os.makedirs(documents_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        document_number = data.get("document_number") or data.get("documentNumber")
        if not document_number:
            return Response({"error": "Document number is required"}, status=400)
        
        excel_filename = f'delivery_challan_{document_number}_{timestamp}.xlsx'
        excel_path = os.path.join(documents_dir, excel_filename)
        
        wb.save(excel_path)
        document_path = f'documents/{excel_filename}'
        excel_url = f'/media/{document_path}'
        
        # Store delivery challan in database
        with transaction.atomic():
            challan = DeliveryChallan.objects.create(
                document_number=document_number,  # Use the validated document number
                project_code=project_code_obj,
                date=data.get('date') or datetime.now().date(),
                reference_no=data.get('reference_no'),
                mode_of_transport=data.get('mode_of_transport'),
                vehicle_no=data.get('vehicle_no'),
                dispatch_from=data.get('dispatch_from'),
                place_of_supply=data.get('place_of_supply'),
                bill_to=data.get('bill_to', ''),
                ship_to=data.get('ship_to', ''),
                remarks=data.get('remarks', ''),
                document_path=document_path,
                cgst=float(data.get('cgst', 0) or 0),
                sgst=float(data.get('sgst', 0) or 0),
                igst=float(data.get('igst', 0) or 0),
                total_amount=final_amount,
                created_by=data.get('created_by')
            )
            
            # Create challan items
            for item_data in challan_items:
                try:
                    inventory = Inventory.objects.get(id=item_data['inventory_id'])
                    DeliveryChallanItem.objects.create(
                        challan=challan,
                        inventory=inventory,
                        item_no=item_data['item_no'],
                        description=item_data['description'],
                        make=item_data['make'],
                        material_group=item_data['material_group'],
                        hsn_code=item_data['hsn_code'],
                        quantity=item_data['quantity'],
                        rate=item_data['rate'],
                        amount=item_data['amount'],
                        uom=item_data['uom']
                    )
                except Inventory.DoesNotExist:
                    logger.warning(f"Inventory with ID {item_data['inventory_id']} not found.")
        
        return Response({
            'message': 'Document generated successfully',
            'download_url': excel_url,
            'file_name': excel_filename,
            'document_number': document_number,
            'document_path': document_path
        }, status=201)

    except Exception as e:
        logger.error(f"Error generating document: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({
            'error': 'Failed to generate document',
            'details': str(e)
        }, status=500)

def number_to_words(number):
    try:
        from num2words import num2words
        return num2words(number, lang='en_IN').title()
    except ImportError:
        return str(number)  # Fallback if num2words is not installed

@api_view(['GET'])
def get_location_stock(request, inventory_id):
    try:
        inventory = Inventory.objects.get(id=inventory_id)
        
        location_mapping = {
            'times_sq_stock': 'Times Square',
            'i_sq_stock': 'iSquare',
            'sakar_stock': 'Sakar',
            'pirana_stock': 'Pirana',
            'other_stock': 'Other'
        }

        location_stocks = {}
        for db_location, display_name in location_mapping.items():
            stock_details = inventory.get_location_stock_details(db_location)
            
            # Get allocations for this location
            allocations = LocationWiseAllocation.objects.filter(
                stock_allocation__inventory=inventory,
                stock_allocation__status='allocated',
                location=db_location
            ).select_related(
                'stock_allocation',
                'stock_allocation__project_code'
            )

            location_stocks[display_name] = {
                'total': float(stock_details['total']),
                'allocated': float(stock_details['allocated']),
                'available': float(stock_details['available']),
                'outward': float(stock_details['outward']),
                'allocations': [{
                    'project_code': alloc.stock_allocation.project_code.code.project_code if alloc.stock_allocation.project_code and alloc.stock_allocation.project_code.code else None,
                    'quantity': float(alloc.quantity),
                    'allocation_date': alloc.stock_allocation.allocation_date,
                    'remarks': alloc.stock_allocation.remarks
                } for alloc in allocations]
            }

        return Response(location_stocks)

    except Inventory.DoesNotExist:
        return Response(
            {'error': 'Inventory not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error in get_location_stock: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_item_project_requirements(request, item_no):
    try:
        projects = Project.objects.filter(
            requisitions__cimcon_part_number=item_no,
            requisitions__approved_status=True
        ).distinct()
        
        requirements = []
        for project in projects:
            requisition_details = Requisition.objects.filter(
                project=project,
                cimcon_part_number=item_no,
                approved_status=True
            ).aggregate(
                total_required=Sum('req_qty'),
                earliest_required_date=Min('required_by_date')
            )
            
            required_qty = requisition_details['total_required'] or 0
            required_by_date = requisition_details['earliest_required_date']
            
            allocated_qty = StockAllocation.objects.filter(
                inventory__item_no=item_no,
                project_code__code=project.project_code
            ).aggregate(
                total_allocated=Sum('allocated_quantity')
            )['total_allocated'] or 0
            
            days_remaining = (required_by_date - timezone.now().date()).days if required_by_date else None
            is_critical = days_remaining is not None and days_remaining < 7
            
            requirements.append({
                'project_code': project.project_code,
                'client_project_name': project.client_project_name,
                'required_quantity': required_qty,
                'allocated_quantity': allocated_qty,
                'pending_quantity': max(0, required_qty - allocated_qty),
                'required_by_date': required_by_date,
                'days_remaining': days_remaining,
                'is_critical': is_critical,
                'priority_level': 'High' if is_critical else (
                    'Medium' if days_remaining and days_remaining < 15 else 'Low'
                )
            })

        requirements.sort(key=lambda x: (0 if x['is_critical'] else 1, x['required_by_date'] or datetime.max.date()))
        
        return Response(requirements)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def debug_location_allocations(request, allocation_id=None):
    """
    Debug endpoint to view LocationWiseAllocation records
    """
    try:
        if allocation_id:
            allocations = LocationWiseAllocation.objects.filter(
                stock_allocation_id=allocation_id
            ).select_related('stock_allocation')
        else:
            allocations = LocationWiseAllocation.objects.all().select_related('stock_allocation')
        
        data = [{
            'id': alloc.id,
            'stock_allocation_id': alloc.stock_allocation_id,
            'location': alloc.location,
            'quantity': float(alloc.quantity),
            'project_code': alloc.stock_allocation.project_code.code.project_code if alloc.stock_allocation.project_code and alloc.stock_allocation.project_code.code else None,
            'inventory_item': alloc.stock_allocation.inventory.item_no
        } for alloc in allocations]
        
        return Response(data)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def validate_project_code(request, project_code):
    """Endpoint to validate if a project code exists"""
    try:
        exists = ProjectCode.objects.filter(code=project_code).exists()
        return Response({'exists': exists})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def check_allocation_availability(request, inventory_id, project_code):
    """
    Check if stock can be allocated to a specific project
    """
    try:
        inventory = Inventory.objects.get(id=inventory_id)
        
        # Get existing allocations
        existing_allocations = StockAllocation.objects.filter(
            inventory=inventory,
            status='allocated'
        ).exclude(
            project_code__code=project_code
        )
        
        # Calculate available stock
        total_allocated = sum(a.allocated_quantity for a in existing_allocations)
        available_for_project = inventory.total_stock - total_allocated
        
        return Response({
            'can_allocate': available_for_project > 0,
            'available_quantity': float(available_for_project),
            'current_allocations': [{
                'project_code': a.project_code.code,
                'quantity': float(a.allocated_quantity)
            } for a in existing_allocations]
        })
        
    except Inventory.DoesNotExist:
        return Response({
            'error': 'Inventory not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
def create_project_code(request):
    try:
        project_code = request.data.get('code')  # This is the project code string
        name = request.data.get('name')
        
        # Get the Project instance
        try:
            project = Project.objects.get(project_code=project_code)
        except Project.DoesNotExist:
            return Response({
                'error': f'Project with code {project_code} does not exist'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Create the project code
        project_code_obj = ProjectCode.objects.create(
            code=project,  # Save the Project instance
            name=name
        )
        
        return Response({
            'id': project_code_obj.id,
            'code': project_code_obj.code.project_code,
            'name': project_code_obj.name
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

class ProjectCodeViewSet(viewsets.ModelViewSet):
    queryset = ProjectCode.objects.all()
    serializer_class = ProjectCodeSerializer

    def perform_create(self, serializer):
        project_code = self.request.data.get('code')
        try:
            project = Project.objects.get(project_code=project_code)
            serializer.save(code=project)
        except Project.DoesNotExist:
            raise serializers.ValidationError(f"Project with code {project_code} does not exist")

def update_po_status(po_number):
    """Helper function to update PO status based on inward quantities"""
    try:
        po = PurchaseOrder.objects.get(po_number=po_number)
        all_po_items = po.line_items.all()
        
        total_items = all_po_items.count()
        completed_items = 0
        partially_inwarded_items = 0
        
        for line_item in all_po_items:
            if line_item.inwarded_quantity >= line_item.quantity:
                completed_items += 1
            elif line_item.inwarded_quantity > 0:
                partially_inwarded_items += 1

        # Determine status
        if completed_items == total_items:
            new_status = 'completed'
        elif completed_items > 0 or partially_inwarded_items > 0:
            new_status = 'partially_inwarded'
        else:
            new_status = 'open'
            
        po.inward_status = new_status
        po.total_inwarded_quantity = sum(
            item.inwarded_quantity or 0 for item in all_po_items
        )
        po.save()
        
        logger.info(f"Updated PO {po_number} status to {new_status} "
                   f"(Completed: {completed_items}/{total_items})")
        
        return new_status
    except Exception as e:
        logger.error(f"Error updating PO status: {str(e)}")
        return None

@api_view(['GET'])
def get_inventory_outward_history(request, inventory_id):
    """
    Get paginated outward history for a specific inventory item
    """
    try:
        # Get query parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        document_type = request.GET.get('document_type')
        status_filter = request.GET.get('status')  # Renamed from 'status'

        # Build base query
        query = StockOutward.objects.filter(inventory_id=inventory_id)

        # Apply filters
        if start_date:
            query = query.filter(outward_date__gte=start_date)
        if end_date:
            query = query.filter(outward_date__lte=end_date)
        if document_type:
            query = query.filter(document_type=document_type)
        if status_filter:  # Use renamed variable
            query = query.filter(status=status_filter)

        # Add select_related for foreign keys to optimize queries
        query = query.select_related(
            'inventory',
            'project_code',
            'stock_allocation'
        )

        # Order by most recent first
        query = query.order_by('-outward_date')

        # Paginate results
        paginator = Paginator(query, page_size)
        page_obj = paginator.get_page(page)

        # Serialize data
        serializer = StockOutwardSerializer(page_obj, many=True)

        # Add summary statistics
        total_quantity = query.aggregate(total=Sum('quantity'))['total'] or 0

        response_data = {
            'results': serializer.data,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'total_records': paginator.count,
            'summary': {
                'total_quantity': float(total_quantity),
                'total_documents': query.values('document_number').distinct().count()
            }
        }

        return Response(response_data)

    except Exception as e:
        logger.error(f"Error fetching outward history: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_project_outward_history(request, project_code):
    """
    Get paginated outward history for a specific project
    """
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        
        # Get date range (default to last 30 days if not specified)
        end_date = request.GET.get('end_date', datetime.now().date())
        start_date = request.GET.get('start_date', end_date - timedelta(days=30))

        query = StockOutward.objects.filter(
            project_code__code=project_code,
            outward_date__range=[start_date, end_date]
        ).select_related(
            'inventory',
            'project_code'
        ).order_by('-outward_date')

        paginator = Paginator(query, page_size)
        page_obj = paginator.get_page(page)
        
        serializer = StockOutwardSerializer(page_obj, many=True)

        return Response({
            'results': serializer.data,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'total_records': paginator.count
        })

    except Exception as e:
        logger.error(f"Error fetching project outward history: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_outward_summary(request):
    """
    Get summary statistics for outward transactions
    """
    try:
        days = int(request.GET.get('days', 30))
        start_date = datetime.now() - timedelta(days=days)

        summary = StockOutward.objects.filter(
            outward_date__gte=start_date
        ).aggregate(
            total_quantity=Sum('quantity'),
            total_documents=models.Count('document_number', distinct=True)
        )

        # Get top projects
        top_projects = StockOutward.objects.filter(
            outward_date__gte=start_date
        ).values(
            'project_code__code'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]

        return Response({
            'summary': {
                'total_quantity': float(summary['total_quantity'] or 0),
                'total_documents': summary['total_documents'] or 0,
            },
            'top_projects': list(top_projects)
        })

    except Exception as e:
        logger.error(f"Error fetching outward summary: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
        
        
@api_view(['POST'])
def preview_document(request):
    try:
        data = request.data
        project_code = data.get('projectCode')
        
        print(f"Received project code: {project_code}")  # Debug log
        
        # Get project details from Project model
        try:
            project = Project.objects.get(project_code=project_code)
            print(f"Found project: {project.project_code}")  # Debug log
            print(f"Bill to: {project.bill_to}")  # Debug log
            print(f"Ship to: {project.ship_to}")  # Debug log
            
            # Get company details from settings or database
            company_details = {
                'name': 'CIMCON SOFTWARE INDIA PRIVATE LIMITED',
                'address': 'CIMCON Software India Pvt. Ltd., Plot No. 1, Sector 16A, Noida - 201301',
                'gstin': '09AABCC1234A1Z5',
                'cin': 'U72200UP2000PTC123456',
                'email': 'info@cimcon.com'
            }

            # Enhanced project details including billing and shipping addresses
            project_details = {
                'code': project.project_code,
                'name': project.client_project_name,
                'bill_to': project.bill_to if project.bill_to else '',
                'ship_to': project.ship_to if project.ship_to else ''
            }

            print(f"Project details: {project_details}")  # Debug log

            # Get items from outward items
            outward_items = data.get('outwardItems', [])
            items = []
            
            for index, item in enumerate(outward_items, 1):
                    # Get inventory details
                    inventory = Inventory.objects.get(id=item['inventory_id'])
                
                    items.append({
                        'sr_no': index,
                        'id': inventory.id,
                        'inventory_id': inventory.id,
                        'item_code': inventory.item_no,
                    'description': item.get('description', ''),
                    'make': item.get('make', ''),
                    'material_group': item.get('material_group', ''),
                    'hsn_code': item.get('hsn_code', ''),
                    'quantity': sum(item.get('location_quantities', {}).values()),
                    'rate': float(item.get('rate', 0)),
                    'amount': float(item.get('rate', 0)) * sum(item.get('location_quantities', {}).values()),
                        'uom': item.get('uom', 'NOS'),
                    'location_quantities': item.get('location_quantities', {})
                })

            # Calculate summary
            summary = {
                'total_amount': sum(item['amount'] for item in items),
                'total_items': len(items),
                'total_quantity': sum(item['quantity'] for item in items)
            }

            preview_data = {
                'company_details': company_details,
                'project_details': project_details,
                'items': items,
                'summary': summary
            }

            return Response(preview_data)

        except Project.DoesNotExist:
            return Response({
                'error': f'Project with code {project_code} not found'
            }, status=404)

    except Exception as e:
        return Response({
            'error': str(e)
        }, status=500)

@api_view(['POST'])
@transaction.atomic
def process_outward_and_document(request):
    """
    Composite endpoint that processes outward stock and generates document
    in a single atomic transaction.
    """
    try:
        # 1. Extract data from request
        data = request.data
        outward_data = data.get('outward', {})
        document_data = data.get('document', {})
        
        logger.info(f"Received outward and document request")
        
        if not outward_data or not document_data:
            return Response(
                {"error": "Both outward and document data are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 2. First, process outward stock using the direct API call
        outward_response = requests.post(
            f"{request.scheme}://{request.get_host()}/outward/",
            json=outward_data,
            headers={'Content-Type': 'application/json'}
        )
        
        if outward_response.status_code != 201:
            # If outward operation failed, return the error
            return Response(
                outward_response.json(),
                status=outward_response.status_code
            )
            
        logger.info(f"Outward processed successfully")
            
        # 3. Then, generate document using the direct API call
        document_response = requests.post(
            f"{request.scheme}://{request.get_host()}/generate-document/",
            json=document_data,
            headers={'Content-Type': 'application/json'}
        )
        
        if document_response.status_code not in [200, 201]:
            # If document generation failed, the transaction will roll back
            # including the outward operation
            logger.error(f"Document generation failed: {document_response.text}")
            try:
                # Try to parse the error response and pass it through
                error_data = document_response.json()
                return Response(
                    error_data,
                    status=document_response.status_code
                )
            except:
                # If can't parse JSON, return the text
                return Response(
                    {"error": document_response.text or "Document generation failed"},
                    status=document_response.status_code
                )
            
        logger.info(f"Document generated successfully")
        
        # 4. Save document data to DeliveryChallan model
        try:
            # Get document response data
            document_result = document_response.json()
            document_path = document_result.get('document_path')
            
            # Get project code instance
            project_code = document_data.get('project_code')
            if not project_code:
                # Try alternate field name if project_code is not found
                project_code = document_data.get('projectCode')
                
            if not project_code:
                return Response({
                    'error': 'Project code is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                project_code_obj = ProjectCode.objects.get(code__project_code=project_code)
            except ProjectCode.DoesNotExist:
                return Response({
                    'error': f'Project code {project_code} not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get document_number from document_data
            document_number = document_data.get('document_number')
            if not document_number:
                return Response({
                    'error': 'Document number is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Create DeliveryChallan
            challan = DeliveryChallan.objects.create(
                document_number=document_number,  # Use the validated document number
                project_code=project_code_obj,
                date=document_data.get('date'),
                reference_no=document_data.get('reference_no'),
                mode_of_transport=document_data.get('mode_of_transport'),
                vehicle_no=document_data.get('vehicle_no'),
                dispatch_from=document_data.get('dispatch_from'),
                place_of_supply=document_data.get('place_of_supply'),
                bill_to=document_data.get('bill_to', ''),
                ship_to=document_data.get('ship_to', ''),
                remarks=document_data.get('remarks', ''),
                document_path=document_path,
                cgst=float(document_data.get('cgst', 0) or 0),
                sgst=float(document_data.get('sgst', 0) or 0),
                igst=float(document_data.get('igst', 0) or 0),
                total_amount=document_data.get('total_amount'),
                created_by=request.user.username if hasattr(request, 'user') else None,
            )
            
            total_amount = 0
            
            # Create DeliveryChallanItem entries
            for item in document_data.get('outward_items', []):
                # Calculate quantity from location quantities
                quantity = float(sum(item.get('location_quantities', {}).values()))
                
                # Get rate and calculate amount
                rate = float(item.get('rate', 0) or 0)
                amount = quantity * rate
                total_amount += amount
                
                # Get inventory instance
                inventory = Inventory.objects.get(id=item.get('inventory_id'))
                
                DeliveryChallanItem.objects.create(
                    challan=challan,
                    inventory=inventory,
                    item_no=item.get('item_no', ''),
                    description=item.get('description', ''),
                    make=item.get('make', ''),
                    material_group=item.get('material_group', ''),
                    hsn_code=item.get('hsn_code', ''),
                    quantity=quantity,
                    rate=rate,
                    amount=amount,
                    uom=item.get('uom', 'NOS')
                )
            
            # Update total amount
            challan.total_amount = total_amount
            challan.save()
            
            logger.info(f"Delivery challan saved with ID: {challan.id}")
            
        except Exception as e:
            logger.error(f"Error saving delivery challan: {str(e)}")
            logger.error(traceback.format_exc())
            # Don't rollback the transaction, as the document was generated successfully
            # Just log the error
            
        # 5. Both operations succeeded, return the document response
        return Response(
            document_response.json(),
            status=document_response.status_code
        )
        
    except Exception as e:
        logger.error(f"Error in process_outward_and_document: {str(e)}")
        logger.error(traceback.format_exc())
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
def create_returnable_gate_pass(request):
    try:
        # Check if we have form data
        if request.FILES and 'document' in request.FILES:
            document_file = request.FILES['document']
            logger.info(f"Document file received: {document_file.name}, size: {document_file.size}")
            
            # Parse JSON data from the form data
            if 'data' in request.POST:
                try:
                    data = json.loads(request.POST.get('data'))
                    logger.info(f"JSON data parsed from form: {data}")
                except Exception as json_error:
                    logger.error(f"Error parsing JSON data: {str(json_error)}")
                    return Response({'error': 'Invalid JSON data'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                data = request.data
        else:
            # No file uploaded
            document_file = None
            data = request.data
            logger.warning("No document file received in request")
            return Response({'error': 'Supporting document is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Creating returnable gate pass with data: {data}")
        
        with transaction.atomic():
            # Handle both FormData and direct JSON
            if 'data' in request.POST:
                # This is a FormData submission with a file
                data = json.loads(request.POST.get('data', '{}'))
            else:
                # This is a regular JSON submission
                data = request.data
            
            # Get file from request if it exists
            document_file = request.FILES.get('document') if request.FILES else None
            if document_file:
                logger.info(f"Document file received: {document_file.name}")
            
            # Check if we have a file in the request
            document_file = None
            if request.FILES and 'document' in request.FILES:
                document_file = request.FILES['document']
                logger.info(f"Document file detected: {document_file.name}")
            
            # Validate required fields based on gate pass type
            pass_type = data.get('pass_type')
            
            required_fields = ['issue_date', 'source_location', 'items']
            if pass_type == 'outward':
                required_fields.extend(['issued_to', 'expected_return_date'])
            elif pass_type == 'internal':
                required_fields.append('destination_location')
            
            for field in required_fields:
                if not data.get(field) and field != 'expected_return_date':  # Special handling for dates
                    return Response({
                        'error': f'Field {field} is required for {pass_type} gate pass'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Special handling for expected_return_date
            expected_return_date = data.get('expected_return_date')
            if pass_type == 'outward' and not expected_return_date:
                return Response({
                    'error': 'Expected return date is required for outward gate pass'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Handle empty strings for dates
            if expected_return_date == '':
                expected_return_date = None
                
            # Generate gate pass number
            today = timezone.now().date()
            year = str(today.year)[-2:]
            month = str(today.month).zfill(2)
            prefix = 'CIMRGP' if pass_type == 'outward' else 'CIMITP'
            
            # Get latest gate pass number
            latest_pass = ReturnableGatePass.objects.filter(
                gate_pass_number__startswith=f"{prefix}-{year}{month}"
            ).order_by('-gate_pass_number').first()
            
            if latest_pass:
                # Extract the sequence number and increment
                seq_num = int(latest_pass.gate_pass_number.split('-')[-1]) + 1
            else:
                seq_num = 1
                
            gate_pass_number = f"{prefix}-{year}{month}-{str(seq_num).zfill(4)}"
            
            # Add this validation before creating the gate pass
            project_code_id = data.get('project_code_id')
            if project_code_id:
                try:
                    project_code_id = int(project_code_id)
                except (ValueError, TypeError):
                    return Response({
                        'error': f"Invalid project_code_id: {project_code_id}. Must be a number."
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Create gate pass with properly processed values
            gate_pass = ReturnableGatePass.objects.create(
                gate_pass_number=gate_pass_number,
                pass_type=pass_type,
                issue_date=data.get('issue_date'),
                expected_return_date=expected_return_date,
                issued_to=data.get('issued_to') or 'Internal Transfer' if pass_type == 'internal' else data.get('issued_to'),
                issued_to_contact=data.get('issued_to_contact'),
                purpose=data.get('purpose', ''),
                source_location=data.get('source_location'),
                destination_location=data.get('destination_location'),
                project_code_id=project_code_id,  # Using validated project_code_id
                remarks=data.get('remarks', ''),
                created_by=data.get('created_by', 'system')
            )
            
            # Handle document file upload
            if document_file:
                try:
                    # Define the directory path
                    media_root = os.path.join(settings.BASE_DIR, 'media')
                    gate_passes_dir = os.path.join(media_root, 'gate_passes')
                    
                    # Create the directory if it doesn't exist (with verbose logging)
                    if not os.path.exists(gate_passes_dir):
                        logger.info(f"Creating directory: {gate_passes_dir}")
                        os.makedirs(gate_passes_dir, exist_ok=True)
                    
                    # Create a filename with the gate pass number
                    file_extension = os.path.splitext(document_file.name)[1]
                    file_name = f"{gate_pass.gate_pass_number}{file_extension}"
                    file_path = os.path.join(gate_passes_dir, file_name)
                    
                    # Log the file path
                    logger.info(f"Saving document to: {file_path}")
                    
                    # Save the file
                    with open(file_path, 'wb+') as destination:
                        for chunk in document_file.chunks():
                            destination.write(chunk)
                    
                    # Update the gate pass record with the file path
                    gate_pass.document_path = f'gate_passes/{file_name}'  # No leading 'media/' or slash
                    gate_pass.save()
                    
                    logger.info(f"Document saved successfully, path: {gate_pass.document_path}")
                except Exception as upload_error:
                    logger.error(f"Error saving document: {str(upload_error)}")
                    logger.error(traceback.format_exc())
            
            # Process items
            for item_data in data.get('items', []):
                inventory_id = item_data.get('inventory_id')
                quantity = Decimal(str(item_data.get('quantity', 0)))
                source_location_field = item_data.get('source_location')
                
                if not inventory_id or quantity <= 0 or not source_location_field:
                    return Response({
                        'error': f'Invalid item data: {item_data}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Get inventory
                try:
                    inventory = Inventory.objects.get(id=inventory_id)
                except Inventory.DoesNotExist:
                    return Response({
                        'error': f'Inventory item with ID {inventory_id} not found'
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Check stock availability
                current_stock = getattr(inventory, source_location_field, 0)
                if current_stock < quantity:
                    return Response({
                        'error': f'Insufficient stock for {inventory.item_no} at {source_location_field}. Available: {current_stock}, Requested: {quantity}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Create gate pass item
                gate_pass_item = ReturnableGatePassItem.objects.create(
                    gate_pass=gate_pass,
                    inventory=inventory,
                    quantity=quantity,
                    source_location=source_location_field,
                    destination_location=item_data.get('destination_location'),
                    condition_on_issue=item_data.get('condition', ''),
                    remarks=item_data.get('remarks', '')
                )
                
                # Update inventory
                if pass_type == 'outward':
                    # For outward, reduce source location stock
                    setattr(inventory, source_location_field, current_stock - quantity)
                    inventory.save()
                else:
                    # For internal transfer, move stock between locations
                    source_stock = getattr(inventory, source_location_field)
                    dest_location = item_data.get('destination_location')
                    
                    if dest_location:
                        setattr(inventory, source_location_field, source_stock - quantity)
                        
                        dest_stock = getattr(inventory, dest_location, 0)
                        setattr(inventory, dest_location, dest_stock + quantity)
                        
                        inventory.save()
            
            # Return success response
            response_data = {
                'success': True,
                'message': 'Gate pass created successfully',
                'gate_pass_number': gate_pass_number,
                'id': gate_pass.id,
                'document_url': gate_pass.document_path if hasattr(gate_pass, 'document_path') and gate_pass.document_path else None
            }
            
            return Response(response_data, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        logger.error(f"Error creating returnable gate pass: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_returnable_gate_passes(request):
    try:
        pass_type = request.query_params.get('type')
        status_filter = request.query_params.get('status')
        
        queryset = ReturnableGatePass.objects.all().order_by('-created_at')
        
        if pass_type:
            queryset = queryset.filter(pass_type=pass_type)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        # Update overdue status for all passes
        for gate_pass in queryset:
            gate_pass.update_status()
            
        # Simple serialization
        result = []
        for gate_pass in queryset:
            items_data = []
            for item in gate_pass.items.all():
                items_data.append({
                    'id': item.id,
                    'item_no': item.inventory.item_no,
                    'description': item.inventory.description,
                    'quantity': float(item.quantity),
                    'returned_quantity': float(item.returned_quantity),
                    'pending_quantity': float(item.quantity - item.returned_quantity),
                    'source_location': item.source_location,
                    'destination_location': item.destination_location
                })
                
            result.append({
                'id': gate_pass.id,
                'gate_pass_number': gate_pass.gate_pass_number,
                'pass_type': gate_pass.pass_type,
                'type_display': gate_pass.get_pass_type_display(),
                'issue_date': gate_pass.issue_date,
                'expected_return_date': gate_pass.expected_return_date,
                'issued_to': gate_pass.issued_to,
                'purpose': gate_pass.purpose,
                'source_location': gate_pass.source_location,
                'destination_location': gate_pass.destination_location,
                'project_code': gate_pass.project_code.code.project_code if gate_pass.project_code and gate_pass.project_code.code else None,
                'status': gate_pass.status,
                'status_display': gate_pass.get_status_display(),
                'created_at': gate_pass.created_at,
                'items': items_data,
                'item_count': len(items_data),
                'total_quantity': sum(item['quantity'] for item in items_data),
                'returned_quantity': sum(item['returned_quantity'] for item in items_data)
            })
            
        return Response(result)
        
    except Exception as e:
        logger.error(f"Error fetching returnable gate passes: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_returnable_gate_pass(request, gate_pass_id):
    try:
        gate_pass = ReturnableGatePass.objects.get(id=gate_pass_id)
        gate_pass.update_status()
        
        # Detailed serialization
        items_data = []
        for item in gate_pass.items.all().select_related('inventory'):
            returns_data = []
            for return_entry in item.return_entries.all().select_related('gate_pass_return'):
                returns_data.append({
                    'id': return_entry.id,
                    'return_date': return_entry.gate_pass_return.return_date,
                    'quantity': float(return_entry.quantity),
                    'condition': return_entry.condition,
                    'remarks': return_entry.remarks
                })
                
            items_data.append({
                'id': item.id,
                'inventory_id': item.inventory.id,
                'item_no': item.inventory.item_no,
                'description': item.inventory.description,
                'make': item.inventory.make,
                'material_group': item.inventory.material_group,
                'quantity': float(item.quantity),
                'returned_quantity': float(item.returned_quantity),
                'pending_quantity': float(item.quantity - item.returned_quantity),
                'source_location': item.source_location,
                'destination_location': item.destination_location,
                'condition_on_issue': item.condition_on_issue,
                'condition_on_return': item.condition_on_return,
                'remarks': item.remarks,
                'returns': returns_data
            })
            
        # Get all return entries
        returns_data = []
        for return_entry in gate_pass.returns.all():
            return_items = []
            for item in return_entry.items.all().select_related('gate_pass_item__inventory'):
                return_items.append({
                    'id': item.id,
                    'item_no': item.gate_pass_item.inventory.item_no,
                    'description': item.gate_pass_item.inventory.description,
                    'quantity': float(item.quantity),
                    'condition': item.condition,
                    'remarks': item.remarks
                })
                
            returns_data.append({
                'id': return_entry.id,
                'return_date': return_entry.return_date,
                'received_by': return_entry.received_by,
                'remarks': return_entry.remarks,
                'items': return_items
            })
            
        result = {
            'id': gate_pass.id,
            'gate_pass_number': gate_pass.gate_pass_number,
            'pass_type': gate_pass.pass_type,
            'type_display': gate_pass.get_pass_type_display(),
            'issue_date': gate_pass.issue_date,
            'expected_return_date': gate_pass.expected_return_date,
            'issued_to': gate_pass.issued_to,
            'issued_to_contact': gate_pass.issued_to_contact,
            'purpose': gate_pass.purpose,
            'source_location': gate_pass.source_location,
            'destination_location': gate_pass.destination_location,
            'project_code': gate_pass.project_code.code.project_code if gate_pass.project_code and gate_pass.project_code.code else None,
            'remarks': gate_pass.remarks,
            'status': gate_pass.status,
            'status_display': gate_pass.get_status_display(),
            'created_by': gate_pass.created_by,
            'created_at': gate_pass.created_at,
            'updated_at': gate_pass.updated_at,
            'items': items_data,
            'returns': returns_data,
            'document_path': gate_pass.document_path if hasattr(gate_pass, 'document_path') else None,
        }
        
        return Response(result)
        
    except ReturnableGatePass.DoesNotExist:
        return Response({
            'error': f'Gate pass with ID {gate_pass_id} not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching returnable gate pass: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def process_gate_pass_return(request, gate_pass_id):
    try:
        data = request.data
        logger.info(f"Processing return for gate pass {gate_pass_id}: {data}")
        
        with transaction.atomic():
            # Get gate pass
            try:
                gate_pass = ReturnableGatePass.objects.get(id=gate_pass_id)
            except ReturnableGatePass.DoesNotExist:
                return Response({
                    'error': f'Gate pass with ID {gate_pass_id} not found'
                }, status=status.HTTP_404_NOT_FOUND)
                
            # Validate gate pass is not fully returned
            if gate_pass.status == 'fully_returned':
                return Response({
                    'error': 'Gate pass is already fully returned'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            # Validate return items
            return_items = data.get('items', [])
            if not return_items:
                return Response({
                    'error': 'No items provided for return'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            # Create return record
            return_record = ReturnableGatePassReturn.objects.create(
                gate_pass=gate_pass,
                return_date=data.get('return_date', timezone.now().date()),
                received_by=data.get('received_by', ''),
                remarks=data.get('remarks', '')
            )
            
            # Process each return item
            for return_item in return_items:
                gate_pass_item_id = return_item.get('gate_pass_item_id')
                quantity = Decimal(str(return_item.get('quantity', 0)))
                
                if not gate_pass_item_id or quantity <= 0:
                    return Response({
                        'error': f'Invalid return item data: {return_item}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                    
                # Get gate pass item
                try:
                    gate_pass_item = ReturnableGatePassItem.objects.get(id=gate_pass_item_id, gate_pass=gate_pass)
                except ReturnableGatePassItem.DoesNotExist:
                    return Response({
                        'error': f'Gate pass item with ID {gate_pass_item_id} not found'
                    }, status=status.HTTP_404_NOT_FOUND)
                    
                # Validate quantity doesn't exceed pending quantity
                pending_qty = gate_pass_item.quantity - gate_pass_item.returned_quantity
                if quantity > pending_qty:
                    return Response({
                        'error': f'Return quantity ({quantity}) exceeds pending quantity ({pending_qty}) for item {gate_pass_item.inventory.item_no}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                    
                # Create return item
                return_item_record = ReturnableGatePassReturnItem.objects.create(
                    gate_pass_return=return_record,
                    gate_pass_item=gate_pass_item,
                    quantity=quantity,
                    condition=return_item.get('condition', ''),
                    remarks=return_item.get('remarks', '')
                )
                
                # Update gate pass item
                gate_pass_item.returned_quantity += quantity
                gate_pass_item.condition_on_return = return_item.get('condition', '')
                gate_pass_item.save()
                
                # Update inventory - increase stock at the specified location
                inventory = gate_pass_item.inventory
                
                # For outward returns, add back to the original source location
                if gate_pass.pass_type == 'outward':
                    source_location_field = gate_pass_item.source_location
                    current_stock = getattr(inventory, source_location_field, 0)
                    setattr(inventory, source_location_field, current_stock + quantity)
                    inventory.save()
                
            # Update gate pass status
            gate_pass.update_status()
            
            return Response({
                'success': True,
                'message': 'Return processed successfully',
                'return_id': return_record.id,
                'gate_pass_status': gate_pass.status
            }, status=status.HTTP_200_OK)
            
    except Exception as e:
        logger.error(f"Error processing gate pass return: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def generate_gate_pass_document(request, gate_pass_id):
    try:
        gate_pass = ReturnableGatePass.objects.get(id=gate_pass_id)
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Gate Pass"
        
        # Document title
        if gate_pass.pass_type == 'outward':
            title = "RETURNABLE GATE PASS"
        else:
            title = "INTERNAL TRANSFER MEMO"
            
        # Add company header
        ws.merge_cells('A1:H1')
        ws['A1'] = "CIMCON SOFTWARE INDIA PRIVATE LIMITED"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add document title
        ws.merge_cells('A2:H2')
        ws['A2'] = title
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Style for headers
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        
        # Gate pass details
        ws['A4'] = "Gate Pass No:"
        ws['B4'] = gate_pass.gate_pass_number
        ws['B4'].font = Font(bold=True)
        
        ws['D4'] = "Date:"
        ws['E4'] = gate_pass.issue_date.strftime('%d/%m/%Y')
        
        if gate_pass.pass_type == 'outward':
            ws['A5'] = "Issued To:"
            ws['B5'] = gate_pass.issued_to
            
            ws['D5'] = "Contact:"
            ws['E5'] = gate_pass.issued_to_contact or 'N/A'
            
            ws['A6'] = "Expected Return Date:"
            ws['B6'] = gate_pass.expected_return_date.strftime('%d/%m/%Y') if gate_pass.expected_return_date else 'N/A'
        else:
            ws['A5'] = "Source Location:"
            ws['B5'] = gate_pass.source_location
            
            ws['D5'] = "Destination Location:"
            ws['E5'] = gate_pass.destination_location or 'N/A'
        
        ws['A7'] = "Purpose:"
        ws['B7'] = gate_pass.purpose
        
        # Add items table headers
        headers = ['S.No.', 'Item Code', 'Description', 'Make', 'Quantity', 'Unit', 'Source', 'Destination' if gate_pass.pass_type == 'internal' else 'Remarks']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
        # Add items data
        row_idx = 10
        for idx, item in enumerate(gate_pass.items.all().select_related('inventory'), 1):
            ws.cell(row=row_idx, column=1).value = idx
            ws.cell(row=row_idx, column=2).value = item.inventory.item_no
            ws.cell(row=row_idx, column=3).value = item.inventory.description
            ws.cell(row=row_idx, column=4).value = item.inventory.make
            ws.cell(row=row_idx, column=5).value = float(item.quantity)
            ws.cell(row=row_idx, column=6).value = 'Nos'
            ws.cell(row=row_idx, column=7).value = item.source_location.replace('_stock', '').title()
            
            if gate_pass.pass_type == 'internal':
                ws.cell(row=row_idx, column=8).value = item.destination_location.replace('_stock', '').title() if item.destination_location else 'N/A'
            else:
                ws.cell(row=row_idx, column=8).value = item.remarks or ''
                
            # Add borders to cells
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
            row_idx += 1
            
        # Add signature sections
        row_idx += 2
        ws.cell(row=row_idx, column=1).value = "Authorized By:"
        ws.cell(row=row_idx, column=5).value = "Issued By:"
        
        if gate_pass.pass_type == 'outward':
            ws.cell(row=row_idx + 5, column=1).value = "Received By:"
        else:
            ws.cell(row=row_idx + 5, column=1).value = "Received At Destination:"
            
        # Save the file
        media_root = os.path.join(settings.BASE_DIR, 'media')
        documents_dir = os.path.join(media_root, 'gate_passes')
        os.makedirs(documents_dir, exist_ok=True)
        
        file_name = f"{gate_pass.gate_pass_number}.xlsx"
        file_path = os.path.join(documents_dir, file_name)
        wb.save(file_path)
        
        # Return the file URL
        file_url = f'/media/gate_passes/{file_name}'
        
        return Response({
            'success': True,
            'file_url': file_url,
            'file_name': file_name
        })
        
    except ReturnableGatePass.DoesNotExist:
        return Response({
            'error': f'Gate pass with ID {gate_pass_id} not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error generating gate pass document: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_inventory_returnable_info(request, inventory_id):
    try:
        inventory = Inventory.objects.get(id=inventory_id)
        
        # Get all gate pass items for this inventory
        gatepass_items = ReturnableGatePassItem.objects.filter(
            inventory=inventory
        ).select_related('gate_pass')
        
        result = []
        for item in gatepass_items:
            gate_pass = item.gate_pass
            result.append({
                'gate_pass_number': gate_pass.gate_pass_number,
                'pass_type': gate_pass.pass_type,
                'issue_date': gate_pass.issue_date,
                'expected_return_date': gate_pass.expected_return_date,
                'status': gate_pass.status,
                'quantity': float(item.quantity),
                'returned_quantity': float(item.returned_quantity),
                'pending_quantity': float(item.quantity - item.returned_quantity)
            })
        
        return Response(result)
    except Inventory.DoesNotExist:
        return Response({'error': 'Inventory not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def get_gate_pass_document(request, gate_pass_id):
    try:
        gate_pass = ReturnableGatePass.objects.get(id=gate_pass_id)
        
        if not gate_pass.document_path:
            return Response({
                'error': 'No document available for this gate pass'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get the filename from document_path
        filename = os.path.basename(gate_pass.document_path)
        
        # Build path directly using settings.MEDIA_ROOT and the document path
        document_path = os.path.join(settings.MEDIA_ROOT, gate_pass.document_path)
        
        # Log the document path for debugging
        logger.info(f"Attempting to access document at: {document_path}")
        
        # Check if the file exists
        if not os.path.exists(document_path):
            return Response({
                'error': f'Document file not found at {document_path}'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get file extension and content type
        file_ext = os.path.splitext(filename)[1].lower()
        
        # Map common extensions to MIME types
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.txt': 'text/plain'
        }
        
        content_type = content_types.get(file_ext, 'application/octet-stream')
        
        # Create a FileResponse with the correct content type
        response = FileResponse(open(document_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except ReturnableGatePass.DoesNotExist:
        return Response({
            'error': f'Gate pass with ID {gate_pass_id} not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error serving document: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_rejected_material_return(request):
    try:
        # Handle file upload similar to returnable gate pass
        document_file = request.FILES.get('document') if request.FILES else None
        
        if 'data' in request.POST:
            try:
                data = json.loads(request.POST.get('data'))
            except Exception as json_error:
                logger.error(f"Error parsing JSON data: {str(json_error)}")
                return Response({'error': 'Invalid JSON data'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            data = request.data
        
        logger.info(f"Creating rejected material return with data: {data}")
        
        with transaction.atomic():
            # Validate required fields
            required_fields = ['challan_number', 'client_name', 'return_date', 'reason_for_return', 'items']
            for field in required_fields:
                if not data.get(field):
                    return Response({
                        'error': f'Field {field} is required'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get project code if specified
            project_code_id = data.get('project_code_id')
            if project_code_id:
                try:
                    project_code_id = int(project_code_id)
                except (ValueError, TypeError):
                    return Response({
                        'error': f"Invalid project_code_id: {project_code_id}. Must be a number."
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create return record
            return_record = RejectedMaterialReturn.objects.create(
                challan_number=data.get('challan_number'),
                client_name=data.get('client_name'),
                return_date=data.get('return_date'),
                project_code_id=project_code_id,
                reason_for_return=data.get('reason_for_return'),
                remarks=data.get('remarks', ''),
                created_by=data.get('created_by', 'system')
            )
            
            # Handle document file upload
            if document_file:
                try:
                    # Create directory if it doesn't exist
                    media_root = os.path.join(settings.BASE_DIR, 'media')
                    returns_dir = os.path.join(media_root, 'returns')
                    os.makedirs(returns_dir, exist_ok=True)
                    
                    # Save file with unique name
                    file_extension = os.path.splitext(document_file.name)[1]
                    file_name = f"return_{return_record.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"
                    file_path = os.path.join(returns_dir, file_name)
                    
                    with open(file_path, 'wb+') as destination:
                        for chunk in document_file.chunks():
                            destination.write(chunk)
                    
                    # Update return record with document path
                    return_record.document_path = f'returns/{file_name}'
                    return_record.save()
                    
                except Exception as upload_error:
                    logger.error(f"Error saving document: {str(upload_error)}")
                    logger.error(traceback.format_exc())
            
            # Process items
            added_to_stock = False
            for item_data in data.get('items', []):
                inventory_id = item_data.get('inventory_id')
                quantity = Decimal(str(item_data.get('quantity', 0)))
                action = item_data.get('action')
                
                if not inventory_id or quantity <= 0 or not action:
                    return Response({
                        'error': f'Invalid item data: {item_data}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Get inventory
                try:
                    inventory = Inventory.objects.get(id=inventory_id)
                except Inventory.DoesNotExist:
                    return Response({
                        'error': f'Inventory item with ID {inventory_id} not found'
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Create return item
                location = item_data.get('location') if action == 'add_to_stock' else None
                
                return_item = RejectedMaterialItem.objects.create(
                    material_return=return_record,
                    inventory=inventory,
                    quantity=quantity,
                    location=location,
                    condition=item_data.get('condition', ''),
                    action=action,
                    reason_details=item_data.get('reason_details', '')
                )
                
                # If adding to stock, update inventory
                if action == 'add_to_stock' and location:
                    added_to_stock = True
                    current_stock = getattr(inventory, location, 0)
                    setattr(inventory, location, current_stock + quantity)
                    inventory.save()
            
            # Update return status based on items
            if added_to_stock:
                return_record.action_taken = 'added_to_stock'
            else:
                return_record.action_taken = 'discarded'
            return_record.save()
            
            return Response({
                'success': True,
                'message': 'Rejected material return processed successfully',
                'id': return_record.id
            }, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        logger.error(f"Error creating rejected material return: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_rejected_material_returns(request):
    try:
        status_filter = request.query_params.get('status')
        
        queryset = RejectedMaterialReturn.objects.all().order_by('-created_at')
        
        if status_filter:
            queryset = queryset.filter(action_taken=status_filter)
            
        # Serialize data
        result = []
        for return_record in queryset:
            items_data = []
            for item in return_record.items.all():
                items_data.append({
                    'id': item.id,
                    'item_no': item.inventory.item_no,
                    'description': item.inventory.description,
                    'quantity': float(item.quantity),
                    'location': item.location,
                    'condition': item.condition,
                    'action': item.action,
                    'reason_details': item.reason_details
                })
                
            result.append({
                'id': return_record.id,
                'challan_number': return_record.challan_number,
                'client_name': return_record.client_name,
                'return_date': return_record.return_date,
                'project_code': return_record.project_code.code.project_code if return_record.project_code and return_record.project_code.code else None,
                'reason_for_return': return_record.reason_for_return,
                'action_taken': return_record.action_taken,
                'status_display': dict(return_record.RETURN_STATUS_CHOICES).get(return_record.action_taken),
                'remarks': return_record.remarks,
                'created_at': return_record.created_at,
                'items': items_data,
                'item_count': len(items_data),
                'has_document': bool(return_record.document_path)
            })
            
        return Response(result)
        
    except Exception as e:
        logger.error(f"Error fetching rejected material returns: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_rejected_material_return(request, return_id):
    try:
        return_record = RejectedMaterialReturn.objects.get(id=return_id)
        
        # Detailed serialization
        items_data = []
        for item in return_record.items.all().select_related('inventory'):
            items_data.append({
                'id': item.id,
                'inventory_id': item.inventory.id,
                'item_no': item.inventory.item_no,
                'description': item.inventory.description,
                'make': item.inventory.make,
                'material_group': item.inventory.material_group,
                'quantity': float(item.quantity),
                'location': item.location,
                'condition': item.condition,
                'action': item.action,
                'reason_details': item.reason_details
            })
            
        result = {
            'id': return_record.id,
            'challan_number': return_record.challan_number,
            'client_name': return_record.client_name,
            'return_date': return_record.return_date,
            'project_code': return_record.project_code.code.project_code if return_record.project_code and return_record.project_code.code else None,
            'project_code_id': return_record.project_code_id,
            'reason_for_return': return_record.reason_for_return,
            'action_taken': return_record.action_taken,
            'status_display': dict(return_record.RETURN_STATUS_CHOICES).get(return_record.action_taken),
            'remarks': return_record.remarks,
            'created_by': return_record.created_by,
            'created_at': return_record.created_at,
            'updated_at': return_record.updated_at,
            'items': items_data,
            'document_path': return_record.document_path
        }
        
        return Response(result)
        
    except RejectedMaterialReturn.DoesNotExist:
        return Response({
            'error': f'Return record with ID {return_id} not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching rejected material return: {str(e)}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_rejected_material_document(request, return_id):
    try:
        return_record = RejectedMaterialReturn.objects.get(id=return_id)
        
        if not return_record.document_path:
            return Response({
                'error': 'No document available for this return record'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get the filename from document_path
        filename = os.path.basename(return_record.document_path)
        
        # Build path directly using settings.MEDIA_ROOT
        document_path = os.path.join(settings.MEDIA_ROOT, return_record.document_path)
        
        # Check if the file exists
        if not os.path.exists(document_path):
            return Response({
                'error': f'Document file not found at {document_path}'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get file extension and content type
        file_ext = os.path.splitext(filename)[1].lower()
        
        # Map common extensions to MIME types
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.txt': 'text/plain'
        }
        
        content_type = content_types.get(file_ext, 'application/octet-stream')
        
        # Create a FileResponse with the correct content type
        response = FileResponse(open(document_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except RejectedMaterialReturn.DoesNotExist:
        return Response({
            'error': f'Return record with ID {return_id} not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error serving document: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_delivery_challans(request):
    """
    Get a list of delivery challans for the rejected material return form
    """
    try:
        # Get recent challans, limited to 100
        challans = DeliveryChallan.objects.all().order_by('-created_at')[:100]
        
        result = []
        for challan in challans:
            # Get basic challan info
            challan_data = {
                'id': challan.id,
                'document_number': challan.document_number,
                'date': challan.date,
                'project_code': {
                    'id': challan.project_code.id,
                    'code': challan.project_code.code.project_code if challan.project_code.code else None,
                    'name': challan.project_code.name
                } if challan.project_code else None,
                'bill_to': challan.bill_to,
                'ship_to': challan.ship_to,
                'items': []
            }
            
            # Get challan items
            for item in challan.items.all():
                challan_data['items'].append({
                    'id': item.id,
                    'inventory_id': item.inventory.id,
                    'item_no': item.item_no,
                    'description': item.description,
                    'make': item.make,
                    'material_group': item.material_group,
                    'quantity': float(item.quantity),
                    'rate': float(item.rate)
                })
            
            result.append(challan_data)
            
        return Response(result)
    except Exception as e:
        logger.error(f"Error fetching delivery challans: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_inventory_rejected_materials(request, inventory_id):
    try:
        # Find all rejected material items for this inventory
        rejected_items = RejectedMaterialItem.objects.filter(
            inventory_id=inventory_id
        ).select_related('material_return', 'inventory')
        
        result = []
        for item in rejected_items:
            return_record = item.material_return
            result.append({
                'id': return_record.id,
                'challan_number': return_record.challan_number,
                'client_name': return_record.client_name,
                'return_date': return_record.return_date,
                'reason_for_return': return_record.reason_for_return,
                'action_taken': return_record.action_taken,
                'quantity': float(item.quantity),
                'condition': item.condition,
                'action': item.action,
                'reason_details': item.reason_details or ""
            })
            
        return Response(result)
    except Exception as e:
        logger.error(f"Error fetching rejected materials for inventory {inventory_id}: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_inventory(request):
    try:
        # OPTIMIZATION: Add caching for inventory data (5 minutes)
        cache_key = 'inventory_list_data'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)
        
        # OPTIMIZATION: Use annotations to reduce database queries
        inventory_items = Inventory.objects.all().annotate(
            outward_stock=Sum('stockoutward__quantity')
        ).values(
            'id',
            'item_no',
            'description',
            'make',
            'material_group',
            'opening_stock',
            'allocated_stock',
            'available_stock',
            'times_sq_stock',
            'i_sq_stock',
            'sakar_stock',
            'pirana_stock',
            'other_stock',
            'total_stock',
            'outward_stock'
        )
        
        # Convert Decimal to float for JSON serialization
        processed_items = []
        for item in inventory_items:
            processed_item = {}
            for key, value in item.items():
                if isinstance(value, Decimal):
                    processed_item[key] = float(value or 0)
                else:
                    processed_item[key] = value
            processed_items.append(processed_item)

        # Cache the processed data for 5 minutes
        cache.set(cache_key, processed_items, 300)

        return Response(processed_items, status=status.HTTP_200_OK)
    except Exception as e:
        print("Error fetching inventory:", str(e))
        return Response(
            {'error': 'Failed to fetch inventory data.'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_all_invoices(request):
    """
    Get all invoices with filtering and pagination
    """
    try:
        # Get query parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 25))
        search = request.GET.get('search', '')
        project_code = request.GET.get('project_code', '')
        vendor = request.GET.get('vendor', '')
        po_number = request.GET.get('po_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        sort_field = request.GET.get('sort_field', 'invoice_date')
        sort_direction = request.GET.get('sort_direction', 'desc')

        # Build query
        queryset = InwardEntry.objects.filter(
            purchase_invoice__isnull=False,
            invoice_number__isnull=False
        ).select_related()

        # Apply filters
        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search) |
                Q(po_number__icontains=search) |
                Q(description__icontains=search)
            )

        if project_code:
            # Get PO details to filter by project code
            po_numbers = PurchaseOrder.objects.filter(
                project_code=project_code
            ).values_list('po_number', flat=True)
            queryset = queryset.filter(po_number__in=po_numbers)

        if vendor:
            # Get PO details to filter by vendor
            po_numbers = PurchaseOrder.objects.filter(
                vendor_name__icontains=vendor
            ).values_list('po_number', flat=True)
            queryset = queryset.filter(po_number__in=po_numbers)

        if po_number:
            queryset = queryset.filter(po_number=po_number)

        if start_date:
            queryset = queryset.filter(invoice_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(invoice_date__lte=end_date)

        # Apply sorting
        if sort_direction == 'desc':
            sort_field = f'-{sort_field}'
        queryset = queryset.order_by(sort_field)

        # Pagination
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)

        # Prepare response data
        invoices = []
        for entry in page_obj:
            # Get PO details for additional information
            try:
                po = PurchaseOrder.objects.get(po_number=entry.po_number)
                vendor_name = po.vendor_name
                project_code = po.project_code
                po_date = po.po_date
                vendor_address = po.vendor_address
                vendor_email = po.vendor_email
                vendor_contact = po.vendor_contact
                vendor_gstin = po.vendor_gstin
            except PurchaseOrder.DoesNotExist:
                vendor_name = "Unknown"
                project_code = "Unknown"
                po_date = None
                vendor_address = None
                vendor_email = None
                vendor_contact = None
                vendor_gstin = None

            invoice_data = {
                'id': entry.id,
                'po_number': entry.po_number,
                'invoice_number': entry.invoice_number,
                'invoice_date': entry.invoice_date.isoformat() if entry.invoice_date else None,
                'vendor_name': vendor_name,
                'project_code': project_code,
                'total_amount': float(entry.ordered_quantity * entry.quantity_received) if entry.ordered_quantity and entry.quantity_received else 0,
                'purchase_invoice': entry.purchase_invoice.url if entry.purchase_invoice else None,
                'created_at': entry.received_date.isoformat() if entry.received_date else None,
                'created_by': 'System',  # You can add user tracking if needed
                'inward_entry_id': entry.id,
                # Additional PO details
                'po_date': po_date.isoformat() if po_date else None,
                'vendor_address': vendor_address,
                'vendor_email': vendor_email,
                'vendor_contact': vendor_contact,
                'vendor_gstin': vendor_gstin,
            }
            invoices.append(invoice_data)

        return Response({
            'results': invoices,
            'count': paginator.count,
            'next': page_obj.has_next(),
            'previous': page_obj.has_previous(),
            'page_size': page_size,
            'current_page': page,
            'total_pages': paginator.num_pages,
        })

    except Exception as e:
        logger.error(f"Error fetching invoices: {str(e)}")
        return Response(
            {'error': 'Failed to fetch invoices'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_po_invoices(request, po_number):
    """
    Get all invoices for a specific PO
    """
    try:
        invoices = InwardEntry.objects.filter(
            po_number=po_number,
            purchase_invoice__isnull=False,
            invoice_number__isnull=False
        ).order_by('-invoice_date')

        invoice_list = []
        for entry in invoices:
            invoice_data = {
                'id': entry.id,
                'invoice_number': entry.invoice_number,
                'invoice_date': entry.invoice_date.isoformat() if entry.invoice_date else None,
                'total_amount': float(entry.ordered_quantity * entry.quantity_received) if entry.ordered_quantity and entry.quantity_received else 0,
                'purchase_invoice': entry.purchase_invoice.url if entry.purchase_invoice else None,
                'created_at': entry.received_date.isoformat() if entry.received_date else None,
                'created_by': 'System',
            }
            invoice_list.append(invoice_data)

        return Response(invoice_list)

    except Exception as e:
        logger.error(f"Error fetching invoices for PO {po_number}: {str(e)}")
        return Response(
            {'error': 'Failed to fetch invoices'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def download_invoice(request, invoice_id):
    """
    Download a specific invoice file
    """
    try:
        inward_entry = InwardEntry.objects.get(id=invoice_id)
        
        if not inward_entry.purchase_invoice:
            return Response(
                {'error': 'Invoice file not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )

        # Get the file path
        file_path = inward_entry.purchase_invoice.path
        
        # Check if file exists
        if not os.path.exists(file_path):
            return Response(
                {'error': 'Invoice file not found on server'}, 
                status=status.HTTP_404_NOT_FOUND
            )

        # Open and return the file
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{inward_entry.invoice_number}_{inward_entry.po_number}.pdf"'
            return response

    except InwardEntry.DoesNotExist:
        return Response(
            {'error': 'Invoice not found'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error downloading invoice {invoice_id}: {str(e)}")
        return Response(
            {'error': 'Failed to download invoice'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
def get_invoice_statistics(request):
    """
    Get invoice statistics for dashboard
    """
    try:
        # Total invoices
        total_invoices = InwardEntry.objects.filter(
            purchase_invoice__isnull=False,
            invoice_number__isnull=False
        ).count()

        # Invoices this month
        from datetime import datetime, timedelta
        first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        invoices_this_month = InwardEntry.objects.filter(
            purchase_invoice__isnull=False,
            invoice_number__isnull=False,
            invoice_date__gte=first_day
        ).count()

        # Total amount
        total_amount = InwardEntry.objects.filter(
            purchase_invoice__isnull=False,
            invoice_number__isnull=False
        ).aggregate(
            total=models.Sum(models.F('ordered_quantity') * models.F('quantity_received'))
        )['total'] or 0

        # Invoices by vendor (top 5)
        vendor_stats = InwardEntry.objects.filter(
            purchase_invoice__isnull=False,
            invoice_number__isnull=False
        ).values('po_number').annotate(
            vendor_name=models.Subquery(
                PurchaseOrder.objects.filter(
                    po_number=models.OuterRef('po_number')
                ).values('vendor_name')[:1]
            )
        ).values('vendor_name').annotate(
            count=models.Count('id'),
            total_amount=models.Sum(models.F('ordered_quantity') * models.F('quantity_received'))
        ).order_by('-total_amount')[:5]

        return Response({
            'total_invoices': total_invoices,
            'invoices_this_month': invoices_this_month,
            'total_amount': float(total_amount),
            'top_vendors': list(vendor_stats),
        })

    except Exception as e:
        logger.error(f"Error fetching invoice statistics: {str(e)}")
        return Response(
            {'error': 'Failed to fetch invoice statistics'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
